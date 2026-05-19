"""Build a touch page from a CSV / TSV / Excel sheet + a directory of icons.

CSV/TSV column schema (header row required, case-insensitive):

    index            int   — button slot (0-based) on the page.
    display_name     str   — label shown for the button.
    icon             str   — filename inside --images dir (or absolute path). Optional.
    kind             str   — one of:
                                launch-app        → creates a macroCommand
                                keyboard          → creates a profileAction
                                macro-keysequence → creates a macroCommand running a keyboard sequence
                                existing-action   → reuses an existing action ID (provide --action_id)
                                empty             → reserve the slot but leave unbound
    action_id        str   — for kind=existing-action: the full $...___... ID to bind.
    exe              str   — for kind=launch-app: executable path.
    args             str   — for kind=launch-app: arguments string (optional).
    cwd              str   — for kind=launch-app: working dir (optional).
    keys             str   — for kind=keyboard: a single combo e.g. "Windows+Shift+S".
                              for kind=macro-keysequence: pipe-separated combos
                              e.g. "Ctrl+A|Delete|H|I".
    color            str   — optional hex RGB (e.g. "#FF8800") — writes to ActionColors.
    fn_action_id     str   — optional fn (modifier) action ID to bind to the button.

Usage:
    python build_page.py <profile_dir> \\
        --sheet buttons.csv \\
        --images ./icons \\
        --workspace 0 \\
        --page-name "My new page" \\
        [--replace-named "Old page"] \\
        [--dry-run]

Excel (.xlsx) is supported when `openpyxl` is installed. CSV/TSV need no deps.
"""
from __future__ import annotations

import argparse
import base64
import csv
import json
import re
import shutil
import sys
from pathlib import Path

from _common import die, find_workspace, load_profile, make_touch_button, new_guid, save_profile
import macros as macros_mod


VALID_KINDS = {"launch-app", "keyboard", "macro-keysequence", "existing-action", "empty"}


# --- sheet readers --------------------------------------------------------

def _read_csv(path: Path, delim: str) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        rdr = csv.DictReader(f, delimiter=delim)
        return [{(k or "").strip().lower(): (v or "").strip() for k, v in row.items()} for row in rdr]


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        die("reading .xlsx requires `pip install openpyxl`, or export your sheet as CSV/TSV")
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    out: list[dict[str, str]] = []
    for row in rows[1:]:
        rec = {}
        for h, v in zip(header, row):
            if not h:
                continue
            rec[h] = "" if v is None else str(v).strip()
        if any(rec.values()):
            out.append(rec)
    return out


def read_sheet(path: Path) -> list[dict[str, str]]:
    ext = path.suffix.lower()
    if ext == ".csv":
        return _read_csv(path, ",")
    if ext == ".tsv":
        return _read_csv(path, "\t")
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(path)
    die(f"unsupported sheet extension: {ext} (use .csv, .tsv, or .xlsx)")
    return []  # unreachable


# --- icon generation ------------------------------------------------------

def _png_to_b64(path: Path) -> str:
    with path.open("rb") as f:
        return base64.b64encode(f.read()).decode("ascii")


def write_ict(out_path: Path, png_path: Path | None, label: str | None) -> None:
    """Write an ActionIcons/<id>.ict composite: full-button image + bottom-strip text."""
    items: list[dict] = []
    if png_path is not None:
        items.append({
            "image": _png_to_b64(png_path),
            "imageFileName": "",
            "imageColor": 4294967295,                # 0xFFFFFFFF
            "imageRotation": "None",
            "isVisible": True,
            "itemType": "Image",
            "area": {"x": 0, "y": 0, "width": 100, "height": 100},
        })
    items.append({
        "text": label or "",
        "originalText": None,
        "textColor": 4294967295,
        "fontSize": 4,
        "fontName": "Arial",
        "isVisible": bool(label),
        "itemType": "Text",
        "area": {"x": 0, "y": 81, "width": 100, "height": 19},
    })
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        json.dump({"backgroundColor": 4278190080, "items": items}, f, indent=4, ensure_ascii=False)


# --- color helpers --------------------------------------------------------

HEX_RE = re.compile(r"^#?([0-9A-Fa-f]{6})$")


def parse_color_to_rgb_int(s: str) -> int:
    m = HEX_RE.match(s.strip())
    if not m:
        die(f"bad color {s!r} — expected #RRGGBB")
    return int(m.group(1), 16)


# --- action creators (delegate to macros.py) ------------------------------

def _ns(**kwargs):
    return type("ns", (), kwargs)()


def make_action_for_row(row: dict[str, str], profile: dict) -> str | None:
    kind = row.get("kind", "").lower()
    name = row.get("display_name", "") or row.get("name", "")
    if kind == "empty":
        return None
    if kind == "existing-action":
        aid = row.get("action_id", "")
        if not aid:
            die(f"row {row}: kind=existing-action requires action_id")
        return aid
    if kind == "launch-app":
        exe = row.get("exe", "")
        if not exe:
            die(f"row {row}: kind=launch-app requires exe")
        ns = _ns(name=name or Path(exe).stem,
                 exe=exe,
                 args_or_empty=row.get("args", ""),
                 cwd_or_empty=row.get("cwd", ""))
        return macros_mod.cmd_add_launch_app(ns, profile)
    if kind == "keyboard":
        keys = row.get("keys", "")
        if not keys:
            die(f"row {row}: kind=keyboard requires keys")
        ns = _ns(name=name or keys, keys=keys)
        return macros_mod.cmd_add_keyboard_action(ns, profile)
    if kind == "macro-keysequence":
        keys_field = row.get("keys", "")
        if not keys_field:
            die(f"row {row}: kind=macro-keysequence requires keys (pipe-separated)")
        ns = _ns(name=name or "Sequence", keys=[k.strip() for k in keys_field.split("|") if k.strip()])
        return macros_mod.cmd_add_macro_keysequence(ns, profile)
    die(f"row {row}: unknown kind {kind!r}. Valid: {sorted(VALID_KINDS)}")


# --- main builder ---------------------------------------------------------

def build_page(profile_dir: Path, sheet_path: Path, images_dir: Path | None,
               workspace_idx: int, page_name: str,
               replace_named: str | None, dry_run: bool) -> None:
    root, profile = load_profile(profile_dir)
    rows = read_sheet(sheet_path)
    if not rows:
        die(f"sheet {sheet_path} is empty")

    # Validate up front so we don't half-build a page.
    indices_seen: set[int] = set()
    for r in rows:
        idx_raw = r.get("index", "").strip()
        if not idx_raw or not idx_raw.lstrip("-").isdigit():
            die(f"row missing valid integer index: {r}")
        idx = int(idx_raw)
        if idx < 0:
            die(f"negative index in row: {r}")
        if idx in indices_seen:
            die(f"duplicate index {idx} in sheet")
        indices_seen.add(idx)
        kind = r.get("kind", "").lower()
        if kind not in VALID_KINDS:
            die(f"row {r}: invalid kind {kind!r}. Valid: {sorted(VALID_KINDS)}")
        if r.get("icon") and images_dir is None:
            die(f"row {r}: icon column set but --images not provided")
        if r.get("icon") and images_dir is not None:
            icon_p = (images_dir / r["icon"]) if not Path(r["icon"]).is_absolute() else Path(r["icon"])
            if not icon_p.is_file():
                die(f"icon file not found: {icon_p}")

    # Figure out page size and prepare the page object.
    mode = profile["layout"]["layoutModes"][0]
    ws = find_workspace(profile, workspace_idx)
    # Infer touch-page slot count from an existing page in this workspace if any, else 15.
    template_slots = 15
    if ws.get("touchPageNames"):
        first = next(p for p in mode["touchPages"] if p["name"] == ws["touchPageNames"][0])
        template_slots = len(first["controls"])
    max_idx = max(indices_seen)
    if max_idx >= template_slots:
        die(f"highest sheet index {max_idx} exceeds page slot count {template_slots} for this device")

    target_page: dict | None = None
    if replace_named:
        target_page = next((p for p in mode["touchPages"] if p["displayName"] == replace_named), None)
        if target_page is None:
            print(f"no existing page named {replace_named!r}; creating new page instead")
    if target_page is None:
        target_page = {
            "$type": "Loupedeck.Service.ProfileLayoutButtonPage, LoupedeckService",
            "name": new_guid(),
            "displayName": page_name,
            "description": None,
            "controls": [make_touch_button() for _ in range(template_slots)],
            "dynamicPageName": None,
            "dynamicPagePluginName": None,
            "dynamicPageNumber": 0,
        }
        mode["touchPages"].append(target_page)
        ws["touchPageNames"].append(target_page["name"])
        print(f"  created touch page {page_name!r} (GUID {target_page['name']}) in workspace {ws['displayName']!r}")
    else:
        # reset controls in-place
        target_page["displayName"] = page_name
        for i in range(len(target_page["controls"])):
            target_page["controls"][i] = make_touch_button()
        print(f"  reset existing touch page {target_page['displayName']!r} (GUID {target_page['name']})")

    color_map_updates: dict[str, int] = {}
    icon_writes: list[tuple[Path, Path | None, str]] = []  # (out_path, png_in, label)

    for row in rows:
        idx = int(row["index"])
        kind = row.get("kind", "").lower()
        if kind == "empty":
            continue

        action_id = make_action_for_row(row, profile)
        if action_id is None:
            continue

        target_page["controls"][idx]["pressAction"] = action_id
        if row.get("fn_action_id"):
            target_page["controls"][idx]["fnPressAction"] = row["fn_action_id"]

        # Carry the row's display_name down to the macro/profileAction we just made (overrides any default the helper set).
        name = row.get("display_name", "")
        if name and kind in ("launch-app", "keyboard", "macro-keysequence"):
            _retitle_action(profile, action_id, name)

        # Stage icon write
        if row.get("icon") and images_dir is not None:
            icon_in = (images_dir / row["icon"]) if not Path(row["icon"]).is_absolute() else Path(row["icon"])
            icon_out = root / "ActionIcons" / f"{action_id}.ict"
            icon_writes.append((icon_out, icon_in, name))
        elif name:
            icon_out = root / "ActionIcons" / f"{action_id}.ict"
            icon_writes.append((icon_out, None, name))

        # Stage color
        if row.get("color"):
            rgb = parse_color_to_rgb_int(row["color"])
            color_map_updates[action_id] = rgb

    # Apply staged writes
    for out_path, png_in, label in icon_writes:
        if dry_run:
            print(f"  [dry-run] would write {out_path.relative_to(root)} "
                  f"(image={'embedded' if png_in else 'none'}, label={label!r})")
        else:
            write_ict(out_path, png_in, label)

    if color_map_updates:
        ac_path = root / "ActionColors" / "ActionColors.json"
        ac_map: dict[str, int] = {}
        if ac_path.is_file():
            with ac_path.open(encoding="utf-8") as f:
                ac_map = json.load(f)
        ac_map.update(color_map_updates)
        if dry_run:
            print(f"  [dry-run] would update {len(color_map_updates)} ActionColors entries")
        else:
            ac_path.parent.mkdir(exist_ok=True)
            with ac_path.open("w", encoding="utf-8") as f:
                json.dump(ac_map, f, indent=4, ensure_ascii=False)

    print(f"\n  bound {len([r for r in rows if r.get('kind') != 'empty'])} button(s) on page {page_name!r}")

    if not dry_run:
        save_profile(root, profile)
        print(f"  saved {root / 'ProfileInfo.json'}")
    else:
        print("  [dry-run] no files written")


def _retitle_action(profile: dict, action_id: str, name: str) -> None:
    """Update the display_name of a freshly-created macro/profileAction so it matches the row."""
    m = re.match(r"\$@Generic___@(Macro|ProfileAction|MacroAdjustment)___([A-F0-9]{32})", action_id)
    if not m:
        return
    kind, guid = m.group(1), m.group(2)
    bucket = {"Macro": "macroCommands", "ProfileAction": "profileActions", "MacroAdjustment": "macroAdjustments"}[kind]
    for e in profile.get(bucket) or []:
        target = action_id if kind == "ProfileAction" else guid
        if e.get("name") == target:
            e["displayName"] = name
            return


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("profile_dir")
    ap.add_argument("--sheet", required=True, help="CSV / TSV / XLSX of button rows")
    ap.add_argument("--images", default=None, help="directory containing the icon PNGs")
    ap.add_argument("--workspace", type=int, default=0)
    ap.add_argument("--page-name", required=True, dest="page_name")
    ap.add_argument("--replace-named", default=None, dest="replace_named",
                    help="reset an existing page with this displayName instead of appending a new one")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    images_dir = Path(args.images) if args.images else None
    if images_dir is not None and not images_dir.is_dir():
        die(f"--images dir not found: {images_dir}")

    build_page(
        profile_dir=Path(args.profile_dir),
        sheet_path=Path(args.sheet),
        images_dir=images_dir,
        workspace_idx=args.workspace,
        page_name=args.page_name,
        replace_named=args.replace_named,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    main()
